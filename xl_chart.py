import os
import pandas as pd
from io import BytesIO
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

class DoclingParser:
    """
    Excel Parsing Wrapper with Chart Data Extraction
    """

    def _extract_charts_from_excel(self, file_obj: BytesIO) -> Dict[int, List[Dict[str, str]]]:
        """
        엑셀 파일(BytesIO)을 받아 차트 정보를 추출합니다.
        
        Returns:
            Dict[PageNum, List[ChartInfo]]
            - PageNum: 1부터 시작하는 시트 번호
            - ChartInfo: 제목, 앞뒤 텍스트, 그리고 '데이터를 변환한 마크다운 표'
        """
        charts_by_page = {}

        try:
            # data_only=True: 수식이 아닌 계산된 값(Value)을 가져오기 위함
            wb = load_workbook(file_obj, data_only=True)

            for page_idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                page_num = page_idx + 1  # 1-based page number
                
                # 시트에 차트가 존재하는지 확인
                # (openpyxl 버전에 따라 sheet.charts 또는 sheet._charts 사용)
                charts = getattr(sheet, "charts", []) or getattr(sheet, "_charts", [])
                
                if not charts:
                    continue

                page_charts = []
                
                for chart in charts:
                    # 1. 차트 제목 추출
                    title = "Untitled Chart"
                    try:
                        if chart.title:
                            # 제목이 직접 입력된 경우
                            if hasattr(chart.title, 'tx') and chart.title.tx.rich:
                                title = chart.title.tx.rich.p[0].r[0].t
                            # 제목이 셀을 참조하는 경우
                            elif hasattr(chart.title, 'tx') and chart.title.tx.strRef:
                                ref_vals = self._get_values_from_ref(wb, chart.title.tx.strRef.f)
                                if ref_vals: title = ref_vals[0]
                    except Exception:
                        pass

                    # 2. 차트 위치 기반 문맥(Pre/Post Text) 추출
                    pre_text, post_text = self._get_chart_context(sheet, chart)

                    # 3. [핵심] 차트 데이터 역추적 및 마크다운 변환
                    # wb(워크북 전체)를 넘겨야 다른 시트 참조도 해결 가능
                    md_table = self._resolve_excel_chart_data(wb, chart)

                    page_charts.append({
                        "title": title,
                        "pre_text": pre_text,
                        "post_text": post_text,
                        "table": md_table
                    })

                if page_charts:
                    charts_by_page[page_num] = page_charts

            return charts_by_page

        except Exception as e:
            print(f"Error extracting charts from Excel: {e}")
            return {}

    def _get_chart_context(self, sheet, chart) -> tuple[str, str]:
        """차트 위/아래에 있는 텍스트를 추출합니다."""
        pre_text = ""
        post_text = ""
        try:
            # 앵커 정보 확인 (TwoCellAnchor 권장)
            anchor = chart.anchor
            if hasattr(anchor, '_from'):
                row_start = anchor._from.row
                col_start = anchor._from.col
                
                # Pre-text: 차트 시작 행 바로 윗줄 확인
                if row_start > 0:
                    # openpyxl cell은 1-based index (row_start는 0-based이므로 +1이 현재행, 윗줄은 그대로 사용 가능하지만 안전하게 계산)
                    cell = sheet.cell(row=row_start, column=col_start + 1)
                    pre_text = str(cell.value) if cell.value else ""
                
                # Post-text: 차트 끝 행 바로 아랫줄 확인
                row_end = anchor.to.row
                cell = sheet.cell(row=row_end + 2, column=col_start + 1)
                post_text = str(cell.value) if cell.value else ""
        except Exception:
            pass # 위치 정보가 없거나 OneCellAnchor인 경우 패스
            
        return pre_text.strip(), post_text.strip()

    def _resolve_excel_chart_data(self, wb, chart) -> str:
        """
        차트 객체가 참조하고 있는 셀 주소(Reference)를 해석하여 
        실제 데이터를 가져온 뒤 Pandas DataFrame -> Markdown으로 변환합니다.
        """
        try:
            data_dict = {}
            categories = []
            
            # --- 1. X축 (카테고리) 데이터 가져오기 ---
            # 보통 첫 번째 시리즈의 카테고리 참조를 공통 X축으로 사용
            if len(chart.series) > 0:
                cat_ref = None
                try:
                    # 문자열 참조 (strRef) 또는 숫자 참조 (numRef) 확인
                    if hasattr(chart.series[0], 'cat'):
                        if chart.series[0].cat and chart.series[0].cat.strRef:
                            cat_ref = chart.series[0].cat.strRef.f
                        elif chart.series[0].cat and chart.series[0].cat.numRef:
                            cat_ref = chart.series[0].cat.numRef.f
                    
                    if cat_ref:
                        categories = self._get_values_from_ref(wb, cat_ref)
                except Exception:
                    pass # 카테고리 추출 실패 시 자동 생성

            # --- 2. Y축 (시리즈 값) 데이터 가져오기 ---
            for series in chart.series:
                # 시리즈 이름
                series_name = "Series"
                if series.title:
                    if hasattr(series.title, 'tx') and series.title.tx.rich:
                        series_name = series.title.tx.rich.p[0].r[0].t
                    elif hasattr(series.title, 'tx') and series.title.tx.strRef:
                        ref_vals = self._get_values_from_ref(wb, series.title.tx.strRef.f)
                        if ref_vals: series_name = str(ref_vals[0])

                # 실제 데이터 값 (Values)
                vals = []
                if series.val:
                    if series.val.numRef:
                        vals = self._get_values_from_ref(wb, series.val.numRef.f)
                
                data_dict[series_name] = vals

            # --- 3. Pandas로 마크다운 변환 ---
            if not data_dict:
                return "(차트 데이터 참조를 찾을 수 없음)"

            # 데이터 길이 맞추기 (가장 긴 데이터 기준)
            max_len = max(len(v) for v in data_dict.values())
            
            # 카테고리가 없거나 길이가 안 맞으면 자동 생성
            if not categories or len(categories) != max_len:
                categories = [f"Item {i+1}" for i in range(max_len)]
            
            # 데이터프레임 생성
            df = pd.DataFrame(data_dict, index=categories[:max_len])
            
            # 마크다운 변환 (float 포맷 등 옵션 조정 가능)
            return df.to_markdown()

        except Exception as e:
            return f"(데이터 파싱 중 오류 발생: {e})"

    def _get_values_from_ref(self, wb, ref_str: str) -> List[Any]:
        """
        'Sheet1!$A$1:$A$5' 형태의 참조 문자열을 받아 실제 값 리스트를 반환합니다.
        """
        try:
            if "!" not in ref_str:
                return []
            
            sheet_part, cell_part = ref_str.split("!")
            # 시트 이름의 작은따옴표 제거 ('Sheet 1' -> Sheet 1)
            sheet_name = sheet_part.replace("'", "")
            
            if sheet_name not in wb.sheetnames:
                return []

            sheet = wb[sheet_name]
            
            # 범위 파싱 (openpyxl 유틸리티 사용)
            min_col, min_row, max_col, max_row = range_boundaries(cell_part)
            
            values = []
            
            # 해당 범위의 셀 값을 순서대로 읽기
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, 
                                       min_col=min_col, max_col=max_col, 
                                       values_only=True):
                for cell in row:
                    # None 값 처리 (빈칸은 0이나 빈 문자열로)
                    values.append(cell if cell is not None else "")
            
            return values
            
        except Exception as e:
            print(f"참조 해석 오류 ({ref_str}): {e}")
            return []

# --- 실행 테스트 ---
if __name__ == "__main__":
    # 테스트할 파일 경로 (여기에 실제 엑셀 파일 경로를 입력하세요)
    INPUT_FILE = "/home/coder/project/yjchoi/docling_parser/temp/chart_example.xlsx"

    if os.path.exists(INPUT_FILE):
        print(f"📂 파일 로딩 중: {INPUT_FILE}")
        
        # 1. 파일 읽기
        with open(INPUT_FILE, "rb") as f:
            file_bytes = BytesIO(f.read())
            
        # 2. 파서 실행
        parser = DoclingParser()
        print("📊 차트 데이터 파싱 시작...")
        
        results = parser._extract_charts_from_excel(file_bytes)
        
        # 3. 결과 출력
        if results:
            for page, charts in results.items():
                print(f"\n[Page {page}] 발견된 차트 {len(charts)}개")
                for i, chart in enumerate(charts):
                    print(f"\n--- Chart {i+1}: {chart['title']} ---")
                    print(f"위치 힌트 (Pre): {chart['pre_text']}")
                    print(f"위치 힌트 (Post): {chart['post_text']}")
                    print("\n[추출된 데이터 표]")
                    print(chart['table'])
        else:
            print("❌ 차트를 찾지 못했거나 에러가 발생했습니다.")
    else:
        print(f"⚠️ 테스트할 파일이 없습니다: {INPUT_FILE}")
        print("경로를 수정하거나 샘플 파일을 준비해주세요.")